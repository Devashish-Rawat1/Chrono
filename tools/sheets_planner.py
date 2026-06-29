"""
Output Stage 2 — styled weekly planner spreadsheet.

Builds a clean weekly planner that visually matches the Student Schedule
template (colored day headers, Century Gothic, bordered grid) but is
constructed FRESH rather than by mutating the template file. Mutating the
template kept breaking its fragile formula-driven time column (the source
of the #REF! errors and "disappearing times after 10:30 PM"), so this
module instead writes a brand-new sheet with full control: explicit time
values, centered task text, and a tidy per-day summary on top.

Layout (top to bottom):
  Row 1 : title  "CHRONO — WEEKLY PLANNER"
  Row 2 : daily window line  "Daily window: 06:00 – 22:30"
  Row 4 : "TASKS"     +  per-day task/hours summary
  Row 5 : "FREE TIME" +  per-day free-hours summary
  Row 7 : grid header  TIME | MON | TUES | ... | SUN
  Row 8+: time grid, one row per interval from wake to sleep, each task
          centered and color-filled, free slots white. Nothing past sleep.
"""

import os
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SHEET_NAME = "Weekly Planner"
INTERVAL_MINUTES = 30
FONT_NAME = "Century Gothic"

# Day name -> grid column (B is TIME, C..I are the 7 days).
_DAY_TO_COLUMN = {
    "Monday": 3, "Tuesday": 4, "Wednesday": 5, "Thursday": 6,
    "Friday": 7, "Saturday": 8, "Sunday": 9,
}
_DAY_ABBR = {
    "Monday": "MON", "Tuesday": "TUES", "Wednesday": "WED", "Thursday": "THURS",
    "Friday": "FRI", "Saturday": "SAT", "Sunday": "SUN",
}
# Day-header fill colors, matching the original template exactly.
_DAY_HEADER_FILL = {
    "Monday": "FF548135", "Tuesday": "FFBF9000", "Wednesday": "FFC55A11",
    "Thursday": "FFBF9000", "Friday": "FF548135", "Saturday": "FF2F5496",
    "Sunday": "FF1F3864",
}
_WEEKDAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

GRID_HEADER_ROW = 8
GRID_FIRST_ROW = 9

# Fixed-meaning fills (ARGB). Mirrors the calendar color scheme.
_TYPE_FILL = {
    "meal": "FFF2C744",       # banana
    "recurring": "FFB7B7B7",  # graphite/grey
    "break": "FFD9EAD3",      # pale sage
    "wake": "FFF6B26B",       # tangerine
    "sleep": "FF6FA8DC",      # blueberry
}
# Per-task fills cycle through these (distinct from the fixed-meaning ones).
_TASK_FILL_CYCLE = [
    "FF9FC5E8",  # light blue
    "FFEA9999",  # light red
    "FFB6D7A8",  # light green
    "FFFFE599",  # light gold
    "FFD5A6BD",  # mauve
    "FFB4A7D6",  # lavender
]

_THIN = Side(style="thin", color="FFBFBFBF")
_GRID_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _parse_hhmm(value: str) -> int:
    h, m = value.split(":")
    return int(h) * 60 + int(m)


def _minutes_to_hhmm(total: int) -> str:
    return f"{total // 60:02d}:{total % 60:02d}"


def _time_obj(minutes: int) -> datetime.time:
    return datetime.time(hour=minutes // 60, minute=minutes % 60)


def _add_minutes(hhmm: str, minutes: int) -> str:
    total = min(_parse_hhmm(hhmm) + minutes, 23 * 60 + 59)
    return _minutes_to_hhmm(total)


def _hours_label(minutes: int) -> str:
    hours = minutes / 60
    return f"{int(hours)}h" if hours == int(hours) else f"{hours:.1f}h"


def _build_task_fill_map(blocks: list) -> dict:
    fill_map, i = {}, 0
    for b in blocks:
        if b.get("type") in ("deep_work", "light_work"):
            name = b.get("label")
            if name not in fill_map:
                fill_map[name] = _TASK_FILL_CYCLE[i % len(_TASK_FILL_CYCLE)]
                i += 1
    return fill_map


def _fill_for_block(block: dict, task_fill_map: dict):
    btype = block.get("type")
    if btype in ("deep_work", "light_work"):
        return task_fill_map.get(block.get("label"))
    return _TYPE_FILL.get(btype)


def _compute_day_stats(blocks: list, wake_minutes: int, sleep_minutes: int) -> dict:
    """Per day: hours per task, and free time (waking minus all scheduled)."""
    waking = sleep_minutes - wake_minutes
    stats = {}
    for day in _WEEKDAY_ORDER:
        day_blocks = [b for b in blocks if b.get("day") == day]
        if not day_blocks:
            continue
        task_minutes, scheduled = {}, 0
        for b in day_blocks:
            try:
                dur = _parse_hhmm(b["end"]) - _parse_hhmm(b["start"])
            except (KeyError, ValueError, AttributeError):
                continue
            scheduled += dur
            if b.get("type") in ("deep_work", "light_work"):
                task_minutes[b["label"]] = task_minutes.get(b["label"], 0) + dur
        stats[day] = {"tasks": task_minutes, "free_minutes": max(0, waking - scheduled)}
    return stats


def fill_planner(
    blocks: list,
    template_path: str,
    output_path: str,
    wake_time: str,
    sleep_time: str = None,
    include_wake_sleep: bool = True,
) -> dict:
    """
    Builds the styled weekly planner and saves to output_path.

    template_path is accepted for API compatibility but no longer needs to
    exist -- the planner is built fresh in the template's visual style, so
    there's no formula chain to break. (If a future caller wants to derive
    styling from the template, this is where it'd hook in.)

    Returns {"output_path", "written", "skipped"}.
    """
    wake_minutes = _parse_hhmm(wake_time)
    sleep_minutes = _parse_hhmm(sleep_time) if sleep_time else wake_minutes + 16 * 60

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False

    # ── Column widths ───────────────────────────────────────────────
    ws.column_dimensions["A"].width = 2          # slim left margin
    ws.column_dimensions["B"].width = 12          # TIME / summary labels
    for col in "CDEFGHI":
        ws.column_dimensions[col].width = 16

    # ── Title + window line ─────────────────────────────────────────
    ws["B1"] = "CHRONO — WEEKLY PLANNER"
    ws["B1"].font = Font(name=FONT_NAME, size=18, bold=True, color="FF595959")
    ws["B2"] = f"Daily window: {wake_time} – {sleep_time}" if sleep_time else f"Wake: {wake_time}"
    ws["B2"].font = Font(name=FONT_NAME, size=10, italic=True, color="FF808080")

    # ── Summary band (rows 4-5): per-day tasks + free time ──────────
    day_stats = _compute_day_stats(blocks, wake_minutes, sleep_minutes)
    _write_summary(ws, day_stats)

    # ── Grid header row (row 7) ─────────────────────────────────────
    hdr_font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFFFF")
    time_hdr = ws.cell(row=GRID_HEADER_ROW, column=2, value="TIME")
    time_hdr.font = hdr_font
    time_hdr.fill = PatternFill("solid", start_color="FF7F7F7F", end_color="FF7F7F7F")
    time_hdr.alignment = _CENTER
    time_hdr.border = _GRID_BORDER
    for day, col in _DAY_TO_COLUMN.items():
        c = ws.cell(row=GRID_HEADER_ROW, column=col, value=_DAY_ABBR[day])
        c.font = hdr_font
        c.fill = PatternFill("solid", start_color=_DAY_HEADER_FILL[day], end_color=_DAY_HEADER_FILL[day])
        c.alignment = _CENTER
        c.border = _GRID_BORDER
    ws.row_dimensions[GRID_HEADER_ROW].height = 26

    # ── Time grid ───────────────────────────────────────────────────
    # One row per interval from wake to sleep INCLUSIVE (so the sleep
    # marker has a row). Times are explicit values -- no formulas, so
    # nothing can break or disappear.
    num_slots = (sleep_minutes - wake_minutes) // INTERVAL_MINUTES + 1
    row_of_minute = {}  # absolute minute -> grid row
    time_label_font = Font(name=FONT_NAME, size=9, color="FF666666")
    time_fill = PatternFill("solid", start_color="FFF3F3F3", end_color="FFF3F3F3")
    white = PatternFill("solid", start_color="FFFFFFFF", end_color="FFFFFFFF")

    for i in range(num_slots):
        minutes = wake_minutes + i * INTERVAL_MINUTES
        row = GRID_FIRST_ROW + i
        row_of_minute[minutes] = row
        ws.row_dimensions[row].height = 24

        tcell = ws.cell(row=row, column=2, value=_time_obj(minutes))
        tcell.number_format = "h:mm AM/PM"
        tcell.font = time_label_font
        tcell.fill = time_fill
        tcell.alignment = Alignment(horizontal="right", vertical="center")
        tcell.border = _GRID_BORDER

        # Default every day cell to white with a border; tasks overwrite below.
        for col in _DAY_TO_COLUMN.values():
            cc = ws.cell(row=row, column=col)
            cc.fill = white
            cc.border = _GRID_BORDER

    last_grid_row = GRID_FIRST_ROW + num_slots - 1

    # ── Place task/event blocks ─────────────────────────────────────
    events = list(blocks)
    if include_wake_sleep and wake_time and sleep_time:
        for day in _DAY_TO_COLUMN:
            events.append({"day": day, "start": wake_time, "end": _add_minutes(wake_time, INTERVAL_MINUTES), "label": "Wake up", "type": "wake"})
            events.append({"day": day, "start": sleep_time, "end": _add_minutes(sleep_time, INTERVAL_MINUTES), "label": "Sleep", "type": "sleep"})

    task_fill_map = _build_task_fill_map(events)
    block_font = Font(name=FONT_NAME, size=9, color="FF1A1A1A")

    written, skipped = 0, 0
    skipped_detail = []

    def snap(minute):
        """Snap a minute to the nearest grid row at/after it, within range."""
        if minute < wake_minutes:
            return None
        off = minute - wake_minutes
        off -= off % INTERVAL_MINUTES
        r = GRID_FIRST_ROW + off // INTERVAL_MINUTES
        return r if r <= last_grid_row else None

    for block in events:
        col = _DAY_TO_COLUMN.get(block.get("day"))
        if col is None:
            skipped += 1; skipped_detail.append({"block": block, "reason": "unknown day"}); continue
        try:
            start_min = _parse_hhmm(block["start"])
            end_min = _parse_hhmm(block["end"])
        except (KeyError, ValueError, AttributeError):
            skipped += 1; skipped_detail.append({"block": block, "reason": "bad time"}); continue

        start_row = snap(start_min)
        if start_row is None:
            skipped += 1; skipped_detail.append({"block": block, "reason": "outside window"}); continue
        end_row = snap(max(start_min, end_min - INTERVAL_MINUTES)) or last_grid_row
        end_row = min(end_row, last_grid_row)

        col_letter = get_column_letter(col)
        if end_row > start_row:
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")

        cell = ws.cell(row=start_row, column=col, value=block.get("label", ""))
        cell.font = block_font
        cell.alignment = _CENTER
        cell.border = _GRID_BORDER
        fill = _fill_for_block(block, task_fill_map)
        if fill:
            cell.fill = PatternFill("solid", start_color=fill, end_color=fill)
        written += 1

    # Freeze header + time column so they stay visible when scrolling.
    ws.freeze_panes = "C8"

    if os.path.dirname(output_path):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    return {"output_path": output_path, "written": written, "skipped": skipped_detail}


def _write_summary(ws, day_stats: dict) -> None:
    """
    Writes the per-day summary in a compact, readable block at rows 4-6:
    day mini-headers (row 4), a 'TASKS' row (5) and a 'FREE TIME' row (6),
    each with one column per day aligned under the grid's day columns.

    Styling matches the user's template: grey label cells with bold white
    text, and light-tinted value cells, all bordered for a clean table.
    """
    label_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
    val_font = Font(name=FONT_NAME, size=9, color="FF333333")
    daycol_font = Font(name=FONT_NAME, size=9, bold=True, color="FF666666")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Fills matching the template: medium grey labels, light value cells.
    label_fill = PatternFill("solid", start_color="FF7F7F7F", end_color="FF7F7F7F")
    value_fill = PatternFill("solid", start_color="FFF2F2F2", end_color="FFF2F2F2")

    days_present = [d for d in _WEEKDAY_ORDER if d in day_stats]

    # Row 4: day mini-headers above the summary values.
    for day in days_present:
        c = ws.cell(row=4, column=_DAY_TO_COLUMN[day], value=_DAY_ABBR[day])
        c.font = daycol_font
        c.alignment = center

    # Row 5: TASKS — list each task with its (per-day) hours.
    lbl5 = ws.cell(row=5, column=2, value="TASKS")
    lbl5.font = label_font
    lbl5.alignment = center
    lbl5.fill = label_fill
    lbl5.border = _GRID_BORDER
    ws.row_dimensions[5].height = 34
    for day in days_present:
        tasks = day_stats[day]["tasks"]
        txt = "\n".join(f"{n} {_hours_label(m)}" for n, m in tasks.items()) or "—"
        c = ws.cell(row=5, column=_DAY_TO_COLUMN[day], value=txt)
        c.font = val_font
        c.alignment = center
        c.fill = value_fill
        c.border = _GRID_BORDER

    # Row 6: FREE TIME — free hours per day.
    lbl6 = ws.cell(row=6, column=2, value="FREE TIME")
    lbl6.font = label_font
    lbl6.alignment = center
    lbl6.fill = label_fill
    lbl6.border = _GRID_BORDER
    ws.row_dimensions[6].height = 22
    for day in days_present:
        c = ws.cell(row=6, column=_DAY_TO_COLUMN[day], value=_hours_label(day_stats[day]["free_minutes"]))
        c.font = val_font
        c.alignment = center
        c.fill = value_fill
        c.border = _GRID_BORDER